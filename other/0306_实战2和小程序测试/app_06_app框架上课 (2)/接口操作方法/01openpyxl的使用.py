"""
============================
Author:柠檬班-木森
Time:2019/11/18
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
import openpyxl


# 第一步： 打开工作簿（读取excel文件中所有的数据保存为工作簿对象）

workbook = openpyxl.load_workbook(r"C:\project\py24_class\py24_14day\project_14day_v1\cases.xlsx")


# 第二步：选中表单对象
sheet = workbook["login"]
print(sheet)

# 第三步：通过表单选中表格读取数据
# 1、读取内容
# 第五行，第四列
# data = sheet.cell(row=5,column=1)
# # print(data.value)


# 2、写入内容
# sheet.cell(row=7,column=3,value='("","")')

# 写入内容之后，一定要保存才会生效
# workbook.save(r"C:\project\py24_class\py24_14day\project_14day_v1\cases.xlsx")

# 3、获取最大行和最大列
# 最大行
# print(sheet.max_row)
# 最大列
# print(sheet.max_column)
# 注意点：不要随便在表格中敲空格

# 4、rows:按行获取所有的格子对象。每一行的格子放入一个元组中
"""
[
(<Cell 'login'.A1>, <Cell 'login'.B1>, <Cell 'login'.C1>, <Cell 'login'.D1>),
(<Cell 'login'.A2>, <Cell 'login'.B2>, <Cell 'login'.C2>, <Cell 'login'.D2>),
(<Cell 'login'.A3>, <Cell 'login'.B3>, <Cell 'login'.C3>, <Cell 'login'.D3>),
(<Cell 'login'.A4>, <Cell 'login'.B4>, <Cell 'login'.C4>, <Cell 'login'.D4>), 
(<Cell 'login'.A5>, <Cell 'login'.B5>, <Cell 'login'.C5>, <Cell 'login'.D5>),
(<Cell 'login'.A6>, <Cell 'login'.B6>, <Cell 'login'.C6>, <Cell 'login'.D6>),
(<Cell 'login'.A7>, <Cell 'login'.B7>, <Cell 'login'.C7>, <Cell 'login'.D7>)
]

"""
print(list(sheet.rows))


